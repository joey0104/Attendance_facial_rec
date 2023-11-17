import numpy as np
import cv2
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import date
from openpyxl.utils import get_column_letter

#getting people
people=[]
folder_path="/Users/joeychiu/Documents/facialRec"
for i in os.listdir(folder_path):
    if(i !=".DS_Store"):
        people.append(i)

#create new excel sheet
wb=Workbook()
ws=wb.active
ws["A1"]="Name"
for i in range (len(people)):
    ws["A"+str(i+2)]=people[i]
wb.save("Attendance.xlsx")


#setup face recognizer
haar_cascade=cv2.CascadeClassifier("haarCasCadeFrontalFace.xml")
face_rec=cv2.face.LBPHFaceRecognizer_create()
face_rec.read('trained_face.yml')

#access camera
cam=cv2.VideoCapture(0)
while True:
    # use set to avoid duplicates
    attended_set=set()
    success, image=cam.read()
    gray_image=cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    detected_face=haar_cascade.detectMultiScale(gray_image, scaleFactor=1.1, minNeighbors=7)
    if len(detected_face)==1:
        for (x,y,w,h) in detected_face:
            face_area = gray_image[y:y + h, x:x + w]
            label=face_rec.predict(gray_image)
            cv2.rectangle(image, (x,y), (x+w, y+h), (254,0,0), thickness=3)
            cv2.putText(image, people[label[0]], (x+100,y-10), cv2.FONT_HERSHEY_PLAIN, 3, (254,0,0), thickness=3)
            attended_set.add(people[label[0]])
    cv2.imshow("Image", image)
    cv2.waitKey(0)

# update attendance excel sheet
next_letter = get_column_letter(ws.max_column+1)
ws[next_letter+"2"]=date.today()
list=list(attended_set)
for person in people:
    if person in list:
        ws[next_letter + str(people.index(person)+2)] = "attended"
    else:
        ws[next_letter + str(people.index(person)+2)] = "absent"
wb.save("Attendance.xlsx")
